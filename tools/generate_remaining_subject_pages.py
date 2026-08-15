from __future__ import annotations

import argparse
import hashlib
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

import generate_subject_academy_pages as legacy
from subject_catalog import BY_SLUG, NEW_CATEGORIES, PROTECTED_CATEGORIES, SUBJECT_CATALOG, SubjectCategory


SITE = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path.home() / "Desktop" / "새 폴더"
PUBLISH_DATE = "2026-08-15"
MODIFIED_DATE = "2026-08-15"
EXPECTED_ROWS = 371
EXPECTED_NEW_DETAILS = 5_194
MASKED_JACCARD_LIMIT = 0.75


SIGNAL_TERMS: dict[str, tuple[tuple[str, tuple[str, ...]], ...]] = {
    "수학": (
        ("개념 연결", ("개념", "원리", "정의")),
        ("계산 정확도", ("계산", "연산", "부호")),
        ("문제 해석", ("문제", "조건", "문장제")),
        ("유형 적용", ("유형", "응용", "적용")),
        ("풀이 기록", ("풀이", "과정", "서술")),
        ("오답 재학습", ("오답", "틀린", "실수")),
        ("시험 점검", ("시험", "내신", "평가")),
        ("학습 습관", ("습관", "계획", "복습")),
    ),
    "영어": (
        ("어휘 회상", ("어휘", "단어", "암기")),
        ("문법 적용", ("문법", "구문", "품사")),
        ("문장 해석", ("해석", "문장", "독해")),
        ("근거 찾기", ("근거", "지문", "내용")),
        ("쓰기 표현", ("쓰기", "영작", "서술")),
        ("오답 재학습", ("오답", "틀린", "실수")),
        ("시험 점검", ("시험", "내신", "평가")),
        ("학습 습관", ("습관", "계획", "복습")),
    ),
}

# Seven explicit, editorially neutral fallbacks per subject.  They are used
# only when the source cell provides no positive keyword evidence (or only one
# evidenced signal); the fallback key is retained for the preflight report and
# is never presented as a claim about the source manuscript.
SIGNAL_FALLBACKS: dict[str, tuple[tuple[str, str], ...]] = {
    "수학": (
        ("개념 연결", "계산 정확도"),
        ("문제 해석", "유형 적용"),
        ("풀이 기록", "오답 재학습"),
        ("시험 점검", "학습 습관"),
        ("개념 연결", "문제 해석"),
        ("계산 정확도", "풀이 기록"),
        ("유형 적용", "시험 점검"),
    ),
    "영어": (
        ("어휘 회상", "문법 적용"),
        ("문장 해석", "근거 찾기"),
        ("쓰기 표현", "오답 재학습"),
        ("시험 점검", "학습 습관"),
        ("어휘 회상", "문장 해석"),
        ("문법 적용", "쓰기 표현"),
        ("근거 찾기", "시험 점검"),
    ),
}

SIGNAL_ROLE_PHRASES = {
    "개념 연결": "개념 사이를 연결하는 과정",
    "계산 정확도": "계산 순서와 정확도",
    "문제 해석": "문제의 조건을 해석하는 과정",
    "유형 적용": "배운 원리를 유형에 적용하는 과정",
    "풀이 기록": "풀이 과정에 남은 학습 흔적",
    "오답 재학습": "오답을 다시 해결하는 과정",
    "시험 점검": "시험을 준비하며 확인할 사항",
    "학습 습관": "한 주의 학습 습관",
    "어휘 회상": "문맥에서 어휘를 떠올리는 과정",
    "문법 적용": "문법을 문장에 적용하는 과정",
    "문장 해석": "문장 구조를 따라 해석하는 과정",
    "근거 찾기": "지문에서 근거를 찾는 과정",
    "쓰기 표현": "생각을 문장으로 표현하는 과정",
}


OPENERS = (
    "상담의 출발점은 진도를 앞당기는 일이 아니라 최근 학습 기록을 읽는 데 있습니다",
    "학원 이름을 비교하기 전에 학생이 혼자 해결하는 구간과 멈추는 구간을 나누어 볼 필요가 있습니다",
    "수업 선택은 교재의 난도보다 설명·연습·확인의 흐름이 학생에게 맞는지 살피는 일에서 시작합니다",
    "최근 시험 한 번의 점수만으로 계획을 정하기보다 풀이 흔적과 복습 간격을 함께 보는 편이 정확합니다",
    "현재 단원을 얼마나 배웠는지와 실제로 다시 설명할 수 있는지는 서로 다를 수 있습니다",
    "상담 전에 학교 진도, 과제 수행 시간, 틀린 문제의 처리 방식을 나누어 적으면 판단이 선명해집니다",
    "무조건 많은 문제보다 학생이 틀린 이유를 찾고 수정하는 과정이 남는 수업인지 확인해야 합니다",
    "학습 계획은 목표 분량보다 매주 점검할 행동을 구체적으로 정할 때 유지하기 쉬워집니다",
    "같은 학년이라도 기초 공백과 문제 해결 속도가 달라 먼저 점검할 순서는 학생마다 달라집니다",
    "등원 횟수만 정하기보다 수업 사이에 복습할 시간까지 포함해 한 주의 학습 리듬을 살펴야 합니다",
)

DIAGNOSIS_ACTIONS: dict[str, tuple[str, ...]] = {
    "수학": (
        "대표 문제를 풀게 한 뒤 계산, 개념 선택, 조건 해석 가운데 처음 흔들린 지점을 기록합니다",
        "정답을 가린 채 풀이 순서를 말로 설명하게 해 이해한 부분과 외운 절차를 구분합니다",
        "최근 오답을 단원별로만 묶지 않고 실수 원인별로 다시 나누어 보완 순서를 정합니다",
        "기본형과 변형 문제의 풀이 시간을 비교해 개념 공백인지 적용 속도의 문제인지 살핍니다",
        "풀이 첫 줄과 검산 과정을 함께 보며 막힘이 시작되는 순간을 구체적으로 찾습니다",
        "비슷한 문제를 간격을 두고 다시 풀어 수정한 방법이 실제로 남았는지 확인합니다",
        "문장제의 조건에 표시한 흔적과 식을 세운 이유를 함께 검토해 해석 과정을 점검합니다",
        "단원별 성취도와 과제 소요 시간을 나란히 놓아 무리 없는 주간 분량을 계산합니다",
    ),
    "영어": (
        "짧은 지문을 읽은 뒤 어휘, 구문, 내용 연결 가운데 해석이 멈춘 지점을 표시합니다",
        "정답의 근거 문장을 직접 찾게 해 감으로 고른 문제와 이해하고 푼 문제를 구분합니다",
        "최근 오답을 어휘, 문법, 독해, 서술형으로 나누어 보완 순서를 정합니다",
        "문장을 끊어 읽는 위치와 핵심 동사를 찾는 과정을 함께 보며 해석 습관을 점검합니다",
        "외운 단어를 문장 안에서 다시 뜻풀이하게 해 회상과 적용의 차이를 확인합니다",
        "유형이 비슷한 문항을 간격을 두고 다시 풀어 근거 찾는 방식이 남았는지 살핍니다",
        "서술형 답안을 쓴 순서를 되짚어 문법 지식이 실제 표현으로 이어지는지 확인합니다",
        "지문 길이별 소요 시간과 정확도를 나란히 놓아 현실적인 주간 분량을 정합니다",
    ),
}

PLAN_ACTIONS: dict[str, tuple[str, ...]] = {
    "수학": (
        "개념을 짧게 설명한 다음 기본형, 변형형, 누적 확인 순으로 문제를 배치합니다",
        "한 번에 많은 유형을 섞기보다 같은 원리가 어떻게 달라지는지 비교하는 묶음으로 연습합니다",
        "오답에는 답만 고치지 않고 처음 선택한 식과 수정한 이유를 함께 남깁니다",
        "과제는 새 문제, 수업 오답, 이전 단원 확인을 나누어 복습 시점이 보이게 구성합니다",
        "풀이 속도를 재기 전 정확한 식 세우기를 안정시키고 이후 제한 시간을 단계적으로 줄입니다",
        "학생 설명을 들은 뒤 교사가 빠진 조건을 되묻는 방식으로 사고 과정을 보완합니다",
        "학교 진도와 누적 공백을 다른 칸에 기록해 시험 준비 중에도 기초 보완이 끊기지 않게 합니다",
        "주간 점검에서는 맞힌 수보다 같은 실수의 빈도가 줄었는지를 먼저 확인합니다",
    ),
    "영어": (
        "어휘를 확인한 뒤 문장 구조, 지문 흐름, 근거 문장 순으로 읽기 단계를 연결합니다",
        "문법 규칙을 외우는 데서 끝내지 않고 교과서 문장과 짧은 영작에 번갈아 적용합니다",
        "독해 오답에는 선택지 판단 근거와 지문 속 근거 위치를 함께 표시합니다",
        "과제는 어휘 회상, 수업 지문 복습, 새 지문 적용을 나누어 학습 간격을 보이게 합니다",
        "해석 속도를 높이기 전에 핵심 동사와 수식 관계를 정확히 찾는 습관을 안정시킵니다",
        "학생이 문장을 다시 설명하면 교사가 빠진 연결어와 논리 관계를 질문으로 보완합니다",
        "학교 진도와 누적 어휘를 따로 기록해 시험 기간에도 기본 학습이 끊기지 않게 합니다",
        "주간 점검에서는 외운 개수보다 문장 안에서 정확히 회상한 비율을 먼저 확인합니다",
    ),
}

FEEDBACKS = (
    "수업 직후의 이해도, 혼자 푼 과제, 다음 시간의 재확인을 한 기록으로 이어 봅니다",
    "완료한 분량과 함께 질문이 생긴 위치를 남겨 다음 수업의 시작점을 구체화합니다",
    "학생에게는 바로 고칠 행동을 한 가지씩 제시하고 학부모에게는 변화 추이를 짧게 공유합니다",
    "매주 같은 기준으로 풀이 시간과 정확도를 기록해 감상이 아닌 변화로 계획을 조정합니다",
    "오답을 고친 날과 다시 확인할 날을 분리해 일회성 해설로 끝나지 않게 합니다",
    "진도표 옆에 이해도와 재학습 여부를 함께 적어 배운 범위와 익힌 범위를 구분합니다",
    "질문 횟수보다 질문 내용이 구체적으로 바뀌는지를 살펴 자기 점검의 성장을 확인합니다",
    "시험 뒤에는 점수만 기록하지 않고 준비 과정에서 유지할 점과 바꿀 점을 나누어 정리합니다",
    "과제 누락이 생기면 양을 늘리기보다 요일별 사용 시간을 다시 배치해 실행 가능성을 높입니다",
    "월간 점검에서는 단원 진도와 누적 공백의 변화를 함께 비교해 다음 우선순위를 정합니다",
)

REGION_NOTES = (
    "이동 시간까지 포함한 귀가 동선이 평일 복습 시간을 지나치게 줄이지 않는지 살펴보세요",
    "학교 일정이 바뀌는 주에도 유지할 수 있는 요일과 시간을 상담에서 구체적으로 맞춰 보세요",
    "가까운 거리뿐 아니라 수업 뒤 질문과 과제 점검이 이어지는 방식까지 함께 비교해 보세요",
    "정규 일정과 시험 기간의 운영이 어떻게 달라지는지 미리 물어 학습 리듬을 계획해 보세요",
    "결석이나 학교 행사 뒤 보완 방식이 있는지 확인하면 장기적인 일정 관리에 도움이 됩니다",
    "등원 전후 자습 시간을 실제 생활표에 넣어 무리 없이 반복할 수 있는지 계산해 보세요",
    "센터 위치와 학교 동선을 함께 보고 늦은 시간대의 이동 계획도 상담에서 확인해 보세요",
    "희망 시간만 전달하기보다 가능한 요일의 우선순위를 정해 두면 상담이 구체적이 됩니다",
)

FOLLOW_THROUGHS = (
    "수업에서 이해한 내용을 집에서 다시 꺼내 보는 날짜를 정하고 실행 여부를 다음 상담 기록에 남깁니다",
    "학생이 혼자 시작한 시간과 질문이 생긴 지점을 짧게 적어 수업 밖 학습의 변화를 확인합니다",
    "한 주 목표를 분량, 정확도, 설명 가능 여부로 나누어 어느 부분이 실제로 달라졌는지 비교합니다",
    "과제를 끝내지 못한 날에는 의지로만 해석하지 않고 분량과 시작 시간부터 다시 조정합니다",
    "학부모 공유 내용은 점수 예측보다 이번 주에 확인한 변화와 다음 점검 행동에 초점을 맞춥니다",
    "학생이 수정한 풀이를 다음 주에 다시 설명하게 해 일시적인 이해와 남은 학습을 구분합니다",
    "교재 진도와 별도로 질문 목록을 관리해 어려움이 줄어드는 순서를 구체적으로 살펴봅니다",
    "시험이 없는 기간에도 누적 확인 문제를 배치해 이전 학습이 유지되는지 일정하게 점검합니다",
    "계획을 바꿀 때에는 학습량 전체를 흔들지 않고 가장 막힌 한 단계부터 조정 이유를 기록합니다",
    "월말에는 학생의 자기 점검 문장과 교사의 관찰을 함께 비교해 다음 목표를 한 가지로 좁힙니다",
)

CONSULTATION_CHOICES = (
    "첫 상담에서 모든 계획을 확정하기보다 1주 동안 확인할 행동과 다시 논의할 날짜를 먼저 정합니다",
    "희망 진도와 현재 수행 수준의 차이를 설명받고 복습을 유지할 범위를 구체적으로 합의합니다",
    "수업 시간표는 등원 가능 여부뿐 아니라 귀가 뒤 과제를 시작할 수 있는 시간까지 고려해 선택합니다",
    "교재 선택 이유와 다음 교재로 넘어가는 기준을 물어 학습 단계가 어떻게 연결되는지 확인합니다",
    "질문하기 어려워하는 학생이라면 교사가 막힌 지점을 발견하는 절차를 상담에서 살펴봅니다",
    "시험 기간의 추가 분량보다 평소 기록을 어떤 방식으로 시험 준비에 연결하는지 먼저 확인합니다",
    "학부모 공유 주기와 전달 항목을 미리 맞춰 과도한 보고보다 필요한 변화가 꾸준히 보이게 합니다",
    "학생의 자신감을 점수와 동일하게 보지 않고 혼자 시작하고 수정하는 행동의 변화를 살펴봅니다",
    "상담 뒤에는 비교 기준을 세 가지 이내로 줄여 학생에게 맞는 운영 방식을 차분히 판단합니다",
    "첫 수업 이후 확인할 질문을 미리 적어 두고 설명 방식과 과제 피드백이 실제로 맞는지 점검합니다",
)

RECOMMENDATIONS: dict[str, tuple[str, ...]] = {
    "수학": (
        "답은 맞지만 풀이 이유를 말로 설명하기 어려운 학생",
        "계산 실수가 반복되어 검산 순서를 만들 필요가 있는 학생",
        "기본 문제와 변형 문제의 연결에서 자주 멈추는 학생",
        "오답을 다시 풀어도 같은 단계에서 막히는 학생",
        "학교 진도와 누적 공백을 함께 관리해야 하는 학생",
        "문장제 조건을 식으로 옮기는 데 시간이 오래 걸리는 학생",
        "과제량보다 주간 학습 순서를 구체화할 필요가 있는 학생",
        "시험 전 몰아서 풀기보다 일정한 복습 간격이 필요한 학생",
    ),
    "영어": (
        "단어는 외웠지만 문장 안에서 뜻을 연결하기 어려운 학생",
        "문법 설명은 알지만 독해와 서술형에 적용하기 어려운 학생",
        "지문을 읽고도 정답의 근거 문장을 찾기 어려운 학생",
        "오답을 다시 풀 때 감으로 선택하는 습관이 남은 학생",
        "학교 진도와 누적 어휘를 함께 관리해야 하는 학생",
        "긴 문장에서 핵심 동사와 수식 관계를 놓치는 학생",
        "과제량보다 주간 학습 순서를 구체화할 필요가 있는 학생",
        "시험 전 몰아서 외우기보다 일정한 회상 간격이 필요한 학생",
    ),
}

CHECKLISTS = (
    "최근 시험지나 단원평가에서 틀린 이유를 세 가지 이상 설명할 수 있는지",
    "학교 진도와 별도로 다시 확인해야 할 이전 단원이 무엇인지",
    "평일에 과제와 복습에 실제로 사용할 수 있는 요일과 시간은 언제인지",
    "질문을 바로 해결하는지, 다음 수업 기록으로 이어 관리하는지",
    "결석이나 학교 행사 뒤 빠진 학습을 보완하는 절차가 있는지",
    "시험 기간에 정규 학습과 내신 준비의 비중을 어떻게 조절하는지",
    "학생의 변화와 다음 주 계획을 어떤 주기로 공유하는지",
    "교재를 정하기 전에 현재 수준을 어떤 자료로 확인하는지",
    "오답을 다시 확인하는 날짜와 통과 기준이 정해져 있는지",
    "수업 시간 외에 혼자 풀 수 있는 분량을 어떻게 산정하는지",
)

FORBIDDEN_AVAILABILITY_PHRASES = (
    "수업 불가", "개설 불가", "개설되지", "지원하지", "미지원", "이용 불가", "불가능",
)

KNOWN_COPY_PATTERNS = {
    "signal_role_duplication": re.compile(r"(?:시험 점검|풀이 기록)\s+(?:기록|점검|점검 결과|확인부터|점검 날짜)"),
    "unsafe_plan_prefix": re.compile(
        r"학습 계획은\s+(?:" + "|".join(
            re.escape(value) for values in PLAN_ACTIONS.values() for value in values
        ) + r")"
    ),
    "unsafe_learning_prefix": re.compile(
        r"학습에서는\s+(?:" + "|".join(
            re.escape(value) for values in PLAN_ACTIONS.values() for value in values
        ) + r")"
    ),
    "unsafe_consultation_prefix": re.compile(
        r"상담을 마친 뒤에는\s+(?:" + "|".join(re.escape(value) for value in CONSULTATION_CHOICES) + r")"
    ),
    "faq_possessive_blocker": re.compile(r"(?:개념 연결|계산 정확도|문제 해석|유형 적용|풀이 기록|오답 재학습|시험 점검|학습 습관|어휘 회상|문법 적용|문장 해석|근거 찾기|쓰기 표현)의 막힌 위치"),
    "feedback_topic_duplication": re.compile(
        r"(?:피드백은|주간 기록은)\s+(?:" + "|".join(re.escape(value) for value in FEEDBACKS) + r")"
    ),
    "faq_role_observation_mismatch": re.compile(
        r"(?:" + "|".join(re.escape(value) for value in SIGNAL_ROLE_PHRASES.values()) + r")(?:을|를) 살피며"
    ),
    "faq_role_instability_mismatch": re.compile(
        r"(?:" + "|".join(re.escape(value) for value in SIGNAL_ROLE_PHRASES.values()) + r")(?:이|가) 흔들리는지도"
    ),
    "checklist_indeterminate_period": re.compile(r"상담 체크 \d+:[^.!?]*인지\."),
    "role_improvement_order_mismatch": re.compile(
        r"(?:" + "|".join(re.escape(value) for value in SIGNAL_ROLE_PHRASES.values()) + r")의 보완 순서와 연결해"
    ),
    "role_change_followup_mismatch": re.compile(
        r"(?:" + "|".join(re.escape(value) for value in SIGNAL_ROLE_PHRASES.values()) + r")의 변화가 다음 확인까지"
    ),
    "faq_process_check_repetition": re.compile(
        r"이 과정에서 .*? 기준으로 어느 단계에서 막히는지, 주간 학습 시간은 어떤지 함께 확인하면"
    ),
    "scenario_role_recheck_repetition": re.compile(
        r"(?:" + "|".join(re.escape(value) for value in SIGNAL_ROLE_PHRASES.values())
        + r")(?:을|를) 다시 (?:해결|확인)할 날짜"
    ),
}


@dataclass(frozen=True, slots=True)
class SourceSignal:
    primary: str
    secondary: str
    source_ok: bool
    primary_positive_evidence: bool
    secondary_positive_evidence: bool
    fallback_key: str | None
    raw: str


@dataclass(frozen=True, slots=True)
class PageCopy:
    title: str
    description: str
    summary: str
    sections: tuple[tuple[str, tuple[str, ...]], ...]
    checklist: tuple[str, ...]
    scenarios: tuple[tuple[str, str], ...]
    faqs: tuple[tuple[str, str], ...]

    def audit_strings(self) -> tuple[str, ...]:
        result: list[str] = [self.summary]
        for heading, paragraphs in self.sections:
            result.append(heading)
            result.extend(paragraphs)
        result.extend(self.checklist)
        for heading, paragraph in self.scenarios:
            result.extend((heading, paragraph))
        for question, answer in self.faqs:
            result.extend((question, answer))
        return tuple(result)


def stable_index(seed: str, namespace: str, size: int) -> int:
    digest = hashlib.sha256(f"{seed}|{namespace}".encode("utf-8")).digest()
    return int.from_bytes(digest[:8], "big") % size


def choose(seed: str, namespace: str, values: Sequence[str]) -> str:
    return values[stable_index(seed, namespace, len(values))]


def distinct_choices(seed: str, namespace: str, values: Sequence[str], count: int) -> tuple[str, ...]:
    ranked = sorted(values, key=lambda value: hashlib.sha256(f"{seed}|{namespace}|{value}".encode("utf-8")).digest())
    return tuple(ranked[:count])


def normalized(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def with_josa(value: str, consonant: str, vowel: str) -> str:
    """Append a Korean particle using the last Hangul syllable in a phrase."""
    last = next((character for character in reversed(value) if "가" <= character <= "힣"), "")
    if not last:
        return value + vowel
    has_batchim = (ord(last) - ord("가")) % 28 != 0
    return value + (consonant if has_batchim else vowel)


def ensure_terminal(value: str, mark: str = ".") -> str:
    value = normalized(value)
    if not value:
        return value
    if re.search(r"[.!?。！？][\"'’”)]*$", value):
        return value
    return value + mark


def ensure_question(value: str) -> str:
    """Normalize a public question so its only terminal mark is a question mark."""
    value = normalized(value)
    if not value:
        return value
    value = re.sub(r"[.!?。！？]+$", "", value).rstrip()
    return value + "?"


def signal_role(label: str) -> str:
    try:
        return SIGNAL_ROLE_PHRASES[label]
    except KeyError as exc:
        raise RuntimeError(f"정의되지 않은 학습 신호 역할: {label}") from exc


def split_values(value: str) -> list[str]:
    result: list[str] = []
    school_or_grade = re.compile(
        r"^(?:[가-힣A-Za-z0-9]+(?:초등학교|중학교|고등학교|여중|여고|초|중|고)|[초중고]\d)$"
    )
    for part in re.split(r"[,/·|\n]+", value or ""):
        part = part.strip()
        if not part:
            continue
        tokens = part.split()
        # A space can be a missing school delimiter ("가람중 나루중"), but
        # is preserved inside names such as "서울 중앙고" unless every token
        # independently has a verified school/grade suffix.
        if len(tokens) > 1 and all(school_or_grade.fullmatch(token) for token in tokens):
            result.extend(tokens)
        else:
            result.append(part)
    return list(dict.fromkeys(result))


def load_centers() -> list[dict[str, str]]:
    rows = legacy.load_centers()
    if len(rows) != EXPECTED_ROWS:
        raise RuntimeError(f"센터 CSV는 {EXPECTED_ROWS}행이어야 합니다: {len(rows)}")
    locals_ = [row.get("근처 수업가능 동네", "").strip() for row in rows]
    if any(not local for local in locals_) or len(locals_) != len(set(locals_)):
        raise RuntimeError("센터 CSV의 동네 값이 비어 있거나 중복됩니다")
    return rows


def load_source_cells(category: SubjectCategory) -> list[str]:
    if not category.workbook:
        raise RuntimeError(f"기존 범주는 신규 원고를 읽지 않습니다: {category.slug}")
    path = SOURCE_DIR / category.workbook
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        cells = [sheet.cell(row=index, column=1).value for index in range(1, EXPECTED_ROWS + 1)]
        trailing = [sheet.cell(row=index, column=1).value for index in range(EXPECTED_ROWS + 1, sheet.max_row + 1)]
    finally:
        workbook.close()
    if any(value not in (None, "") for value in trailing):
        raise RuntimeError(f"{path.name} A열에 371행 이후 원고가 있습니다")
    return ["" if value is None else str(value) for value in cells]


def extract_signal(raw: str, category: SubjectCategory, row_index: int) -> SourceSignal:
    source_ok = bool(raw.strip()) and not raw.lstrip().startswith(("#ERROR", "#REF!", "#VALUE!", "="))
    haystack = re.sub(r"<[^>]+>", " ", raw).lower() if source_ok else ""
    scores: list[tuple[int, str]] = []
    for label, terms in SIGNAL_TERMS[category.subject]:
        scores.append((sum(haystack.count(term.lower()) for term in terms), label))
    seed = f"{category.slug}|{row_index}|signal"
    ranked = sorted(scores, key=lambda item: (-item[0], hashlib.sha256(f"{seed}|{item[1]}".encode()).digest()))
    fallback_index = stable_index(seed, "template-fallback", len(SIGNAL_FALLBACKS[category.subject]))
    fallback_pair = SIGNAL_FALLBACKS[category.subject][fallback_index]
    fallback_key = f"{category.subject}-template-fallback-{fallback_index + 1:02d}"
    if not source_ok or ranked[0][0] == 0:
        return SourceSignal(
            fallback_pair[0], fallback_pair[1], source_ok,
            False, False, fallback_key, raw,
        )
    positive = [label for score, label in ranked if score > 0]
    primary = positive[0]
    if len(positive) >= 2:
        return SourceSignal(primary, positive[1], True, True, True, None, raw)
    secondary = next(label for label in fallback_pair if label != primary)
    if secondary == primary:
        labels = tuple(label for label, _ in SIGNAL_TERMS[category.subject])
        secondary = next(label for label in labels if label != primary)
    return SourceSignal(primary, secondary, True, True, False, fallback_key, raw)


def grade_is_listed(row: dict[str, str], category: SubjectCategory) -> bool:
    values = {re.sub(r"\s+", "", value) for value in split_values(row.get(category.grade_field, ""))}
    return category.grade_code in values


def grade_display(row: dict[str, str], category: SubjectCategory) -> str:
    if not grade_is_listed(row, category):
        return "상담 시 확인"
    values = split_values(row.get(category.grade_field, ""))
    return "·".join(values) if values else category.grade_label


def build_copy(row: dict[str, str], category: SubjectCategory, signal: SourceSignal, row_index: int) -> PageCopy:
    local = row["근처 수업가능 동네"].strip()
    region = row.get("지역", "").strip()
    district = row.get("시or구", "").strip()
    center = f"{local} 센터"
    schools = split_values(row.get(category.school_field, ""))
    title = f"{local} {category.label}"
    seed = f"{category.slug}|{row_index}|{local}"
    listed = grade_is_listed(row, category)
    availability = grade_display(row, category)
    school_note = "·".join(schools[:3]) if schools else "재학 학교"
    description = (
        f"{title}: {region} {district}에서 {category.grade_label} {category.subject} 학습을 검토할 때 필요한 "
        f"진단, 복습, 오답 관리와 상담 확인 항목을 안내합니다."
    )
    summary = (
        f"{title} 상담에서는 {with_josa(signal.primary, '과', '와')} {with_josa(signal.secondary, '을', '를')} 함께 점검하고, "
        f"{school_note} 진도 및 학생의 주간 학습 시간을 바탕으로 우선순위를 정하는 것이 핵심입니다."
    )

    opener = choose(seed, "opener", OPENERS)
    diagnostic = choose(seed, "diagnostic", DIAGNOSIS_ACTIONS[category.subject])
    plans = distinct_choices(seed, "plans", PLAN_ACTIONS[category.subject], 2)
    feedbacks = distinct_choices(seed, "feedbacks", FEEDBACKS, 2)
    follow_throughs = distinct_choices(seed, "follow-throughs", FOLLOW_THROUGHS, 2)
    consultation_choices = distinct_choices(seed, "consultation-choices", CONSULTATION_CHOICES, 2)
    region_note = choose(seed, "region", REGION_NOTES)
    recommended = distinct_choices(seed, "recommended", RECOMMENDATIONS[category.subject], 3)
    status_sentence = (
        f"{title} 상담에서 {center} 안내에 표시된 {category.subject} 가능 학년은 {availability}이며, 실제 반 편성과 시간표는 상담에서 확인합니다."
        if listed else
        f"{title} 상담에서 {center}의 {category.grade_label} {category.subject} 가능 학년 표시는 확인 항목이며, 이 안내는 개별 편성을 단정하지 않습니다."
    )
    school_sentence = (
        f"{title} 상담에서는 {with_josa(school_note, '을', '를')} 학교 참고 정보로 확인하고, 학교별 범위는 최신 진도표와 평가 안내를 기준으로 다시 살펴봅니다."
        if schools else
        f"{title} 상담에서는 재학 학교와 최신 진도표를 먼저 확인해 학교별 학습 범위를 구체화합니다."
    )

    sections: tuple[tuple[str, tuple[str, ...]], ...] = (
        (
            f"{local} {category.label} 핵심 답변",
            (
                f"{title}을 알아볼 때 {opener}. {with_josa(signal.primary, '을', '를')} 중심으로 최근 학습 흔적을 확인하면 첫 계획을 과도하게 잡지 않을 수 있습니다.",
                f"{title}의 {category.grade_label} 상담은 점수 하나보다 {signal.secondary}, 과제 소요 시간, 다시 푼 기록을 함께 볼 때 학생에게 필요한 지원을 구체화할 수 있습니다.",
            ),
        ),
        (
            f"{category.grade_label} {category.subject} 진단에서 찾을 답",
            (
                f"{title}의 초기 점검에서는 {diagnostic}. 그 결과를 바탕으로 {signal.primary}에서 먼저 보완할 부분을 정하고 첫 주의 확인 목표를 세웁니다.",
                f"{title}에서 학생의 답안은 맞고 틀림만 세지 않고 시도한 순서와 멈춘 위치를 함께 봅니다. {with_josa(signal_role(signal.secondary), '은', '는')} 설명과 연습의 비중을 조절하는 근거가 됩니다.",
            ),
        ),
        (
            f"{region} {district} 지역·학년·추천 학생 안내",
            (
                f"{school_sentence} {region_note}",
                f"{title} 안내가 특히 도움이 되는 학생은 {recommended[0]}, {recommended[1]}, {recommended[2]}입니다. 현재 학습 자료를 가져오면 필요한 순서를 더 분명하게 논의할 수 있습니다.",
                status_sentence,
            ),
        ),
        (
            f"{local} {category.subject} 설명·연습·재확인 설계",
            (
                f"{title}의 학습 설계 기준을 구체화합니다. {plans[0]}. 이 과정에서 {with_josa(signal.primary, '과', '와')} {with_josa(signal.secondary, '을', '를')} 같은 주간 기록에서 비교합니다.",
                f"{local}의 {category.grade_label} 학생에게 적용할 다음 단계도 정합니다. {plans[1]}. 학생이 혼자 시도할 구간과 도움을 요청할 구간을 구분하면 수업의 흐름이 분명해집니다.",
            ),
        ),
        (
            f"{local} 학습 기록과 학부모 공유의 답",
            (
                f"{title}에서 학부모와 공유할 관찰 기준을 정합니다. {feedbacks[0]}. {follow_throughs[0]}",
                f"{title}의 주간 변화도 같은 기준으로 확인합니다. {feedbacks[1]}. {follow_throughs[1]}",
            ),
        ),
        (
            f"{title} 상담 뒤 결정할 순서",
            (
                f"{title} 선택에 앞서 비교할 기준을 정합니다. {consultation_choices[0]}. {with_josa(signal_role(signal.primary), '을', '를')} 살펴본 결과를 첫 비교 자료로 삼습니다.",
                f"{title} 선택 내용을 다시 확인할 기준도 마련합니다. {consultation_choices[1]}. {signal.secondary}에서 확인한 변화가 다음 점검까지 이어지는지도 함께 살펴봅니다.",
            ),
        ),
    )

    checklist = tuple(
        f"{title} 상담 체크 {index}: {item}"
        for index, item in enumerate(distinct_choices(seed, "checklist", CHECKLISTS, 5), 1)
    )
    scenarios = (
        (
            f"가상 학부모 상담 사례 1 · {signal.primary}",
            f"{title}을 알아보는 한 {category.grade_label} 학부모가 최근 결과에 비해 공부 시간이 길다고 상담한 상황을 가정했습니다. 답안이 멈춘 위치를 토대로 {signal_role(signal.primary)}부터 살펴보고, 첫 주에는 학생이 스스로 설명할 수 있는 범위를 기록하는 방향을 제안합니다.",
        ),
        (
            f"가상 학부모 상담 사례 2 · {signal.secondary}",
            f"{title}을 찾는 또 다른 가정이 과제는 끝내지만 같은 실수가 반복된다고 문의한 상황을 가정했습니다. {signal.secondary}에서 달라져야 할 부분을 살펴볼 날짜를 따로 두고 다음 수업에서 수정 방법을 다시 말하게 하는 흐름을 상담 예시로 안내합니다.",
        ),
    )
    faq_variants = build_faqs(
        row, category, signal, row_index, title, availability, schools,
        used_plans=plans, used_feedbacks=feedbacks,
    )
    finalized_sections = tuple(
        (normalized(heading), tuple(ensure_terminal(paragraph) for paragraph in paragraphs))
        for heading, paragraphs in sections
    )
    finalized_scenarios = tuple(
        (normalized(heading), ensure_terminal(paragraph)) for heading, paragraph in scenarios
    )
    finalized_faqs = tuple(
        (ensure_question(question), ensure_terminal(answer)) for question, answer in faq_variants
    )
    return PageCopy(
        title,
        ensure_terminal(description),
        ensure_terminal(summary),
        finalized_sections,
        tuple(ensure_question(item) for item in checklist),
        finalized_scenarios,
        finalized_faqs,
    )


def build_faqs(
    row: dict[str, str], category: SubjectCategory, signal: SourceSignal, row_index: int,
    title: str, availability: str, schools: list[str], *,
    used_plans: Sequence[str], used_feedbacks: Sequence[str],
) -> tuple[tuple[str, str], ...]:
    local = row["근처 수업가능 동네"].strip()
    center = f"{local} 센터"
    seed = f"{category.slug}|{row_index}|{local}|faq"
    remaining_plans = tuple(value for value in PLAN_ACTIONS[category.subject] if value not in used_plans)
    remaining_feedbacks = tuple(value for value in FEEDBACKS if value not in used_feedbacks)
    if not remaining_plans or not remaining_feedbacks:
        raise RuntimeError(f"FAQ 전용 행동문 후보가 없습니다: {category.slug}/{row_index + 1}")
    first_action = choose(seed, "faq-action", remaining_plans)
    feedback = choose(seed, "faq-feedback", remaining_feedbacks)
    school_text = "·".join(schools[:3]) if schools else "재학 학교"
    listed = grade_is_listed(row, category)
    availability_answer = (
        f"{center} 안내에 표시된 {category.subject} 가능 학년은 {availability}입니다. {title}의 실제 반 편성, 시간표와 학생 진도는 상담 시 함께 확인하세요."
        if listed else
        f"{center}의 {category.grade_label} {category.subject} 가능 학년 표시는 상담 시 확인합니다. {title} 문의 때 현재 진도와 희망 시간을 전달하면 확인이 구체적입니다."
    )
    return (
        (
            f"{local} {category.grade_label} {category.subject} 상담에는 무엇을 가져가면 좋나요?",
            f"{title} 상담에는 최근 시험지나 단원평가, 풀던 교재, 오답 흔적 중 한두 가지면 충분합니다. 이 자료로 {signal.primary}에서 학생이 막히는 지점과 주간 학습 시간을 함께 파악하면 첫 계획을 구체화할 수 있습니다.",
        ),
        (
            f"{title}에서 복습과 다음 진도는 어떻게 나누나요?",
            f"{title}에서는 학생이 현재 내용을 혼자 설명하고 비슷한 문제에 다시 적용하는지 먼저 봅니다. {signal.secondary} 관련 어려움이 반복되는지도 확인합니다. {first_action}.",
        ),
        (
            f"{local} 학교 진도는 수업 계획에 어떻게 반영하나요?",
            f"{school_text} 등 학교별 범위를 미리 단정하지 않고 학생이 가져온 최신 진도표와 평가 안내를 확인합니다. {title} 계획은 학교 일정과 누적 공백을 분리해 기록한 뒤 주별 비중을 조정합니다.",
        ),
        (
            f"{title}의 과제와 피드백은 무엇을 확인해야 하나요?",
            f"{title} 상담에서는 학생이 실제로 쓸 수 있는 시간에 맞춰 과제량을 정하는지 살펴보세요. {feedback}. 완료 여부와 함께 막힌 이유가 다음 수업에 반영되는지도 중요합니다.",
        ),
        (
            f"{center}의 {category.subject} 가능 학년은 어떻게 확인하나요?",
            availability_answer,
        ),
    )


def representative_images() -> list[Path]:
    images = sorted(path for path in (SITE / "assets" / "representative").iterdir() if path.is_file())
    if len(images) != EXPECTED_ROWS:
        raise RuntimeError(f"대표이미지 풀은 {EXPECTED_ROWS}개여야 합니다: {len(images)}")
    return images


def file_sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def existing_representative_usage(
    rows: list[dict[str, str]], images: list[Path],
) -> tuple[list[set[str]], dict[str, int]]:
    pattern = re.compile(
        r'<img\b(?=[^>]*\bclass="[^"]*subject-hidden-representative[^"]*")(?=[^>]*\bsrc="([^"]+)")[^>]*>',
        re.I,
    )
    pool_root = (SITE / "assets" / "representative").resolve()
    used_sha: list[set[str]] = [set() for _ in rows]
    baseline_path_collisions = 0
    baseline_sha_collisions = 0
    for row_index, row in enumerate(rows):
        paths: list[str] = []
        hashes: list[str] = []
        local = row["근처 수업가능 동네"].strip()
        for category in PROTECTED_CATEGORIES:
            page = SITE / "과목별학원" / category.slug / legacy.slug_local(local) / "index.html"
            if not page.is_file():
                raise FileNotFoundError(f"기존 범주 대표이미지를 읽을 수 없습니다: {page}")
            match = pattern.search(page.read_text(encoding="utf-8"))
            if not match:
                raise RuntimeError(f"기존 범주 대표이미지 태그가 없습니다: {page}")
            src = match.group(1).split("?", 1)[0].split("#", 1)[0]
            asset = (SITE / src.lstrip("/")).resolve()
            try:
                asset.relative_to(pool_root)
            except ValueError as exc:
                raise RuntimeError(f"기존 대표이미지가 371 풀 밖에 있습니다: {src}") from exc
            if not asset.is_file():
                raise FileNotFoundError(asset)
            paths.append(src)
            hashes.append(file_sha256(asset))
        baseline_path_collisions += len(paths) - len(set(paths))
        baseline_sha_collisions += len(hashes) - len(set(hashes))
        used_sha[row_index].update(hashes)
    return used_sha, {
        "preexisting_path_collision_excess": baseline_path_collisions,
        "preexisting_sha_collision_excess": baseline_sha_collisions,
    }


def build_representative_assignments(
    rows: list[dict[str, str]], images: list[Path], categories: Sequence[SubjectCategory],
) -> tuple[dict[str, tuple[str, ...]], dict[str, int]]:
    """Build stable per-category perfect matchings against actual existing use.

    Every new category consumes every path exactly once.  A locality never
    receives an image whose bytes are already used there by an existing or
    earlier new category.  The seven existing-existing collisions are retained
    as an immutable baseline because the four protected trees are byte-frozen.
    """
    used_sha, baseline = existing_representative_usage(rows, images)
    image_sha = tuple(file_sha256(path) for path in images)
    assignments: dict[str, tuple[str, ...]] = {}
    for category in categories:
        choices: list[list[int]] = []
        for row_index, row in enumerate(rows):
            local = row["근처 수업가능 동네"].strip()
            candidates = [index for index, digest in enumerate(image_sha) if digest not in used_sha[row_index]]
            candidates.sort(
                key=lambda index: hashlib.sha256(
                    f"{category.slug}|{local}|{images[index].name}|matching".encode("utf-8")
                ).digest()
            )
            choices.append(candidates)

        image_to_row = [-1] * len(images)
        row_to_image = [-1] * len(rows)

        def augment(row_index: int, seen: set[int]) -> bool:
            for image_index in choices[row_index]:
                if image_index in seen:
                    continue
                seen.add(image_index)
                previous = image_to_row[image_index]
                if previous == -1 or augment(previous, seen):
                    image_to_row[image_index] = row_index
                    row_to_image[row_index] = image_index
                    return True
            return False

        row_order = sorted(
            range(len(rows)),
            key=lambda index: (
                len(choices[index]),
                hashlib.sha256(f"{category.slug}|{rows[index]['근처 수업가능 동네']}|row".encode("utf-8")).digest(),
            ),
        )
        for row_index in row_order:
            if not augment(row_index, set()):
                raise RuntimeError(f"대표이미지 완전매칭 실패: {category.slug}/{row_index + 1}")
        if -1 in row_to_image or len(set(row_to_image)) != EXPECTED_ROWS:
            raise RuntimeError(f"대표이미지 순열 불완전: {category.slug}")
        selected = tuple("/assets/representative/" + images[index].name for index in row_to_image)
        assignments[category.slug] = selected
        for row_index, image_index in enumerate(row_to_image):
            digest = image_sha[image_index]
            if digest in used_sha[row_index]:
                raise RuntimeError(f"신규 대표이미지 SHA 충돌: {category.slug}/{row_index + 1}")
            used_sha[row_index].add(digest)

    if len(assignments) != len(categories):
        raise RuntimeError("대표이미지 범주 배정 수가 다릅니다")
    baseline.update({
        "new_category_permutations": len(assignments),
        "new_path_collisions_added": 0,
        "new_sha_collisions_added": 0,
    })
    return assignments, baseline


def representative_for(
    assignments: dict[str, tuple[str, ...]], category: SubjectCategory, row_index: int,
) -> str:
    return assignments[category.slug][row_index]


def protected_hashes() -> dict[str, str]:
    result: dict[str, str] = {}
    for category in PROTECTED_CATEGORIES:
        root = SITE / "과목별학원" / category.slug
        if not root.is_dir():
            raise FileNotFoundError(f"보호할 기존 범주가 없습니다: {root}")
        for path in sorted(item for item in root.rglob("*") if item.is_file()):
            result[path.relative_to(SITE).as_posix()] = hashlib.sha256(path.read_bytes()).hexdigest()
    return result


def assert_protected_unchanged(before: dict[str, str]) -> None:
    after = protected_hashes()
    if before != after:
        changed = sorted(set(before) ^ set(after) | {key for key in before.keys() & after.keys() if before[key] != after[key]})
        raise RuntimeError(f"기존 4개 범주 byte-preserve 위반: {changed[:8]}")


def new_detail_hashes() -> dict[str, str]:
    """Hash only the 5,194 generated detail documents, excluding category hubs."""
    result: dict[str, str] = {}
    for category in NEW_CATEGORIES:
        root = SITE / "과목별학원" / category.slug
        for path in sorted(root.glob("*/index.html")):
            result[path.relative_to(SITE).as_posix()] = hashlib.sha256(path.read_bytes()).hexdigest()
    if len(result) != EXPECTED_NEW_DETAILS:
        raise RuntimeError(f"신규 상세 보호 범위 불일치: {len(result)} != {EXPECTED_NEW_DETAILS}")
    return result


def assert_new_details_unchanged(before: dict[str, str]) -> None:
    after = new_detail_hashes()
    if before != after:
        changed = sorted(set(before) ^ set(after) | {key for key in before.keys() & after.keys() if before[key] != after[key]})
        raise RuntimeError(f"신규 상세 byte-preserve 위반: {changed[:8]}")


def related_links(rows: list[dict[str, str]], row_index: int, category: SubjectCategory) -> list[tuple[str, str, str]]:
    row = rows[row_index]
    local = row["근처 수업가능 동네"].strip()
    sibling_subject = next(
        item for item in SUBJECT_CATALOG
        if item.level == category.level and item.grade_number == category.grade_number and item.subject != category.subject
    )
    neighbors = [rows[(row_index + step) % len(rows)] for step in (1, 7, 29)]
    links = [
        (f"{category.label} 전체 지역", f"/과목별학원/{category.slug}/", "지역 목록 보기"),
        (f"{local} {sibling_subject.label}", f"/과목별학원/{sibling_subject.slug}/{legacy.slug_local(local)}/", "같은 학년 다른 과목"),
        (f"{local} 지역 학원", f"/전국학원/{category.national_category}/{legacy.slug_local(local)}/", "지역 학원 안내"),
    ]
    links.extend(
        (f"{neighbor['근처 수업가능 동네']} {category.label}", f"/과목별학원/{category.slug}/{legacy.slug_local(neighbor['근처 수업가능 동네'])}/", "다른 지역 안내")
        for neighbor in neighbors
    )
    return links


def local_page(
    row: dict[str, str], row_index: int, category: SubjectCategory, copy: PageCopy,
    signal: SourceSignal, rows: list[dict[str, str]], representative_assignments: dict[str, tuple[str, ...]],
) -> str:
    local = row["근처 수업가능 동네"].strip()
    slug = legacy.slug_local(local)
    region = row.get("지역", "").strip()
    district = row.get("시or구", "").strip()
    center = row.get("센터명", "").strip() or f"{local} 학습코칭센터"
    address = row.get("센터 주소", "").strip()
    location = row.get("위치안내", "").strip()
    schools = split_values(row.get(category.school_field, ""))
    registration = " · ".join(filter(None, (row.get("교육지원청명칭", "").strip(), row.get("교육지원청 등록번호", "").strip())))
    canonical = legacy.absolute(f"/과목별학원/{category.slug}/{slug}/")
    page_id = canonical + "#webpage"
    article_id = canonical + "#article"
    service_id = canonical + "#service"
    faq_id = canonical + "#faq"
    breadcrumb_id = canonical + "#breadcrumb"
    org_id = legacy.center_identity_for_row(row)
    representative = representative_for(representative_assignments, category, row_index)
    center_image = "/assets/centers/common/seoul6839.webp" if region == "서울" else "/assets/centers/common/local6839.webp"
    map_image = legacy.find_map(row)
    grade_value = grade_display(row, category)
    listed = grade_is_listed(row, category)
    links = related_links(rows, row_index, category)
    fee = legacy.fee_link(row)

    offer: dict[str, object] = {
        "@type": "Offer", "url": canonical,
        "eligibleCustomerType": category.grade_label if listed else "상담 시 확인",
        "itemOffered": {"@type": "Service", "name": f"{copy.title} 학습 상담", "serviceType": "TutoringService"},
    }
    org_offer = {
        "@type": "Offer", "url": legacy.absolute("/상담문의/"),
        "itemOffered": {"@type": "Service", "name": "학습 상담", "serviceType": "TutoringService"},
    }
    postal_address: dict[str, str] = {"@type": "PostalAddress", "addressCountry": "KR"}
    if address:
        postal_address["streetAddress"] = address
    if region:
        postal_address["addressRegion"] = region
    if district:
        postal_address["addressLocality"] = district
    org: dict[str, object] = {
        "@type": ["EducationalOrganization", "LocalBusiness"], "@id": org_id, "name": center,
        "alternateName": [legacy.SITE_NAME], "telephone": legacy.PHONE_DISPLAY,
        "image": [legacy.absolute(center_image)], "address": postal_address,
        "areaServed": legacy.stable_center_areas(rows, org_id), "knowsAbout": legacy.stable_center_topics(row),
        "makesOffer": [org_offer],
    }
    if registration:
        org["identifier"] = registration
    related_items = [
        {"@type": "ListItem", "position": index, "name": name, "url": legacy.absolute(url)}
        for index, (name, url, _) in enumerate(links, 1)
    ]
    graph: list[dict[str, object]] = [
        {
            "@type": "WebPage", "@id": page_id, "url": canonical, "name": copy.title,
            "description": copy.description, "inLanguage": "ko-KR", "about": {"@id": org_id},
            "mentions": [category.grade_label, category.subject, signal.primary, signal.secondary, local, *schools],
            "primaryImageOfPage": {"@id": canonical + "#primaryimage"}, "breadcrumb": {"@id": breadcrumb_id},
            "mainEntity": {"@id": article_id},
            "hasPart": [{"@id": article_id}, {"@id": service_id}, {"@id": faq_id}, {"@id": canonical + "#related"}],
        },
        {"@type": "ImageObject", "@id": canonical + "#primaryimage", "url": legacy.absolute(representative), "caption": f"{copy.title} 채움학습 대표"},
        {
            "@type": "BreadcrumbList", "@id": breadcrumb_id,
            "itemListElement": [
                {"@type": "ListItem", "position": 1, "name": "홈", "item": legacy.BASE_URL + "/"},
                {"@type": "ListItem", "position": 2, "name": "과목별학원", "item": legacy.absolute("/과목별학원/")},
                {"@type": "ListItem", "position": 3, "name": category.label, "item": legacy.absolute(f"/과목별학원/{category.slug}/")},
                {"@type": "ListItem", "position": 4, "name": copy.title, "item": canonical},
            ],
        },
        org,
        {
            "@type": "Article", "@id": article_id, "mainEntityOfPage": {"@id": page_id},
            "headline": copy.title, "name": copy.title, "description": copy.summary,
            "image": [legacy.absolute(representative), legacy.absolute(center_image), legacy.absolute(map_image)],
            "datePublished": PUBLISH_DATE, "dateModified": MODIFIED_DATE,
            "author": {"@id": org_id}, "publisher": {"@id": org_id}, "inLanguage": "ko-KR",
            "articleSection": [heading for heading, _ in copy.sections] + [f"{local} 상담 전 체크리스트", f"{local} 가상 학부모 상담 사례"],
            "about": [copy.title, category.grade_label, category.subject, signal.primary, signal.secondary],
            "mentions": [local, district, region, center, *schools],
            "hasPart": [{"@id": service_id}, {"@id": faq_id}],
        },
        {
            "@type": "Service", "@id": service_id, "name": f"{copy.title} 학습관리", "serviceType": "TutoringService",
            "provider": {"@id": org_id}, "description": copy.summary,
            "areaServed": {"@type": "Place", "name": local},
            "audience": {"@type": "EducationalAudience", "educationalRole": "student", "audienceType": category.grade_label if listed else "상담 시 확인"},
            "about": [category.subject, signal.primary, signal.secondary], "mentions": ["학습 진단", "오답 재학습", *schools],
            "offers": [offer],
        },
        {
            "@type": "FAQPage", "@id": faq_id,
            "mainEntity": [
                {"@type": "Question", "name": question, "acceptedAnswer": {"@type": "Answer", "text": answer}}
                for question, answer in copy.faqs
            ],
        },
        {"@type": "ItemList", "@id": canonical + "#related", "name": f"{local} 관련 학원 페이지", "numberOfItems": len(related_items), "itemListElement": related_items},
    ]
    if schools:
        graph.append({
            "@type": "ItemList", "@id": canonical + "#schools", "name": f"{local} 수업 학교 상담 참고",
            "numberOfItems": len(schools),
            "itemListElement": [{"@type": "ListItem", "position": index, "name": school} for index, school in enumerate(schools, 1)],
        })

    section_html = "".join(
        f'<section class="section subject-content-section"><div class="section-head"><h2>{legacy.esc(heading)}</h2></div>'
        + "".join(f"<p>{legacy.esc(paragraph)}</p>" for paragraph in paragraphs) + "</section>"
        for heading, paragraphs in copy.sections
    )
    checklist_html = "".join(f"<li>{legacy.esc(item)}</li>" for item in copy.checklist)
    scenario_html = "".join(
        f'<article class="subject-scenario-card info-card"><span class="tag">가상 사례</span><h3>{legacy.esc(heading)}</h3><p>{legacy.esc(paragraph)}</p></article>'
        for heading, paragraph in copy.scenarios
    )
    faq_html = "".join(
        f'<details class="faq-item"><summary>{legacy.esc(question)}</summary><p>{legacy.esc(answer)}</p></details>'
        for question, answer in copy.faqs
    )
    related_html = "".join(
        f'<a class="subject-related-link" href="{legacy.esc(url)}"><strong>{legacy.esc(name)}</strong><small>{legacy.esc(note)}</small></a>'
        for name, url, note in links
    )
    school_html = "".join(f"<li>{legacy.esc(school)}</li>" for school in schools) or "<li>재학 학교와 최신 진도는 상담 시 확인</li>"
    fee_html = f'<a class="btn btn-ghost" href="{legacy.esc(fee)}" target="_blank" rel="noopener">센터 교습비 확인</a>' if fee else ""
    service_notice = ""
    if not legacy.direct_center_area(local, center, address):
        service_notice = f'<p class="subject-service-area-notice">{legacy.esc(local)} 학생의 상담 가능 여부는 인접 센터 정보를 기준으로 확인하며, 실제 센터 위치는 아래 주소에서 확인해 주세요.</p>'

    body = f'''{legacy.nav("과목별학원")}
  <main>
    <section class="page-hero subject-local-hero">
      <nav class="breadcrumb" aria-label="현재 위치"><a href="/">홈</a><span aria-hidden="true">/</span><a href="/과목별학원/">과목별학원</a><span aria-hidden="true">/</span><a href="/과목별학원/{category.slug}/">{legacy.esc(category.label)}</a><span aria-hidden="true">/</span><span aria-current="page">{legacy.esc(copy.title)}</span></nav>
      <p class="eyebrow">LOCAL SUBJECT ACADEMY</p><h1>{legacy.esc(copy.title)}</h1><p class="lead">{legacy.esc(copy.description)}</p>
      <div class="hero-actions"><a class="btn btn-primary" href="tel:{legacy.PHONE_DISPLAY}">학습 상담하기</a>{fee_html}</div>
    </section>
    <section class="section subject-media-section" aria-label="{legacy.esc(copy.title)} 이미지 안내">
      <img class="subject-hidden-representative" src="{legacy.esc(representative)}" alt="{legacy.esc(copy.title)} 채움학습 대표" style="display:none;">
      <div class="media-row">
        <figure class="frame"><img src="{legacy.esc(center_image)}" alt="{legacy.esc(copy.title)} 채움학습 본문" width="1200" height="900" fetchpriority="high"><figcaption>{legacy.esc(local)} 학습관리 안내</figcaption></figure>
        <figure class="frame"><img src="{legacy.esc(map_image)}" alt="{legacy.esc(copy.title)} 채움학습 지도" width="1200" height="900" loading="lazy"><figcaption>{legacy.esc(center)} 위치 안내</figcaption></figure>
      </div>
    </section>
    <section class="section subject-quick-answer subject-content-section"><div class="section-head"><p class="eyebrow">핵심 요약</p><h2>{legacy.esc(local)} {legacy.esc(category.label)} 한눈에 보기</h2><p class="lead">{legacy.esc(copy.summary)}</p></div>
      <div class="subject-fact-grid"><article><span>대상</span><strong>{legacy.esc(category.grade_label)}</strong><p>가능 학년: {legacy.esc(grade_value)}</p></article><article><span>과목</span><strong>{legacy.esc(category.subject)}</strong><p>{legacy.esc(signal.primary)} · {legacy.esc(signal.secondary)}</p></article><article><span>지역</span><strong>{legacy.esc(local)}</strong><p>{legacy.esc(' · '.join(filter(None, (region, district))))}</p></article></div>
    </section>
    {section_html}
    <section class="section subject-checklist subject-content-section"><div class="section-head"><p class="eyebrow">상담 전 체크리스트</p><h2>{legacy.esc(local)} 상담 전에 적어볼 다섯 가지</h2></div><ol>{checklist_html}</ol></section>
    <section class="section subject-scenarios subject-content-section"><div class="section-head"><p class="eyebrow">가상 학부모 상담</p><h2>{legacy.esc(copy.title)} 상담 상황 예시</h2><p class="lead">아래 {legacy.esc(local)} {legacy.esc(category.grade_label)} {legacy.esc(category.subject)} 내용은 실제 후기가 아닌 가상 상담 예시이며, 학습 결과는 학생의 참여와 학습 환경에 따라 달라질 수 있습니다.</p></div><div class="card-grid">{scenario_html}</div></section>
    <section class="section subject-center-card"><div class="section-head"><p class="eyebrow">센터 정보</p><h2>{legacy.esc(center)}</h2><p class="lead">{legacy.esc(copy.title)} 상담에서 주소, 학교 참고 정보와 등록 내용을 확인할 수 있도록 정리했습니다.</p></div>{service_notice}<dl class="subject-center-facts"><div><dt>주소</dt><dd>{legacy.esc(address or '상담 시 안내')}</dd></div><div><dt>위치 안내</dt><dd>{legacy.esc(location or '상담 시 상세 안내')}</dd></div><div><dt>학교 참고</dt><dd><ul>{school_html}</ul></dd></div><div><dt>등록 정보</dt><dd>{legacy.esc(registration or '센터별 등록 정보는 상담 시 확인')}</dd></div></dl>{fee_html}</section>
    <section class="section"><div class="section-head"><p class="eyebrow">FAQ</p><h2>{legacy.esc(local)} {legacy.esc(category.label)} 자주 묻는 질문</h2></div><div class="faq-list">{faq_html}</div></section>
    <section class="section"><div class="section-head"><p class="eyebrow">내부 안내</p><h2>{legacy.esc(local)}에서 함께 확인할 페이지</h2></div><div class="subject-related-grid">{related_html}</div></section>
  </main>
{legacy.footer()}'''
    return legacy.shell(legacy.head_html(copy.title, copy.description, canonical, representative, graph), body)


def region_directory(rows: list[dict[str, str]], category: SubjectCategory) -> str:
    grouped: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        grouped[row.get("지역", "기타").strip() or "기타"][row.get("시or구", "기타").strip() or "기타"].append(row)
    jump = "".join(
        f'<a href="#region-{index}" data-region-target="region-{index}">{legacy.esc(region)}</a>'
        for index, region in enumerate(grouped)
    )
    blocks: list[str] = []
    for region_index, (region, districts) in enumerate(grouped.items()):
        district_blocks: list[str] = []
        for district, district_rows in districts.items():
            links = "".join(
                f'<a href="/과목별학원/{category.slug}/{legacy.slug_local(row["근처 수업가능 동네"])}/"><strong>{legacy.esc(row["근처 수업가능 동네"])}</strong><small>{legacy.esc(category.label)}</small></a>'
                for row in district_rows
            )
            district_blocks.append(f'<section class="subject-district"><h3>{legacy.esc(district)} <small>{len(district_rows)}곳</small></h3><div class="subject-local-grid">{links}</div></section>')
        count = sum(len(values) for values in districts.values())
        blocks.append(f'<section class="region-block" id="region-{region_index}"><div class="region-title"><h2>{legacy.esc(region)}</h2><span>{count}개 지역</span></div>{"".join(district_blocks)}</section>')
    return f'<nav class="region-jump" aria-label="광역 지역 바로가기">{jump}</nav>{"".join(blocks)}'


DIRECTORY_SEARCH_SCRIPT = r'''<script>
(() => {
  "use strict";
  const search = document.querySelector("[data-subject-search]");
  const directory = document.querySelector(".subject-directory");
  if (!search || !directory) return;
  const input = search.querySelector("[data-local-search-input]");
  const reset = search.querySelector("[data-local-search-reset]");
  const status = search.querySelector("[data-local-search-status]");
  const links = Array.from(directory.querySelectorAll(".subject-local-grid > a"));
  const districts = Array.from(directory.querySelectorAll(".subject-district"));
  const regions = Array.from(directory.querySelectorAll(".region-block"));
  const jumps = Array.from(directory.querySelectorAll(".region-jump [data-region-target]"));
  if (!input || !reset || !status || links.length !== 371) return;
  const normalize = value => value.toLocaleLowerCase("ko-KR").replace(/\s+/g, "");
  const update = () => {
    const query = normalize(input.value);
    let count = 0;
    links.forEach(link => {
      const matches = !query || normalize(link.textContent || "").includes(query);
      link.hidden = !matches;
      if (matches) count += 1;
    });
    districts.forEach(district => {
      district.hidden = !district.querySelector(".subject-local-grid > a:not([hidden])");
    });
    regions.forEach(region => {
      region.hidden = !region.querySelector(".subject-local-grid > a:not([hidden])");
    });
    jumps.forEach(jump => {
      const target = document.getElementById(jump.dataset.regionTarget || "");
      jump.hidden = !target || target.hidden;
    });
    status.textContent = query ? `검색 결과 ${count}개 지역` : `전체 ${count}개 지역`;
    reset.disabled = !input.value;
  };
  input.addEventListener("input", update);
  reset.addEventListener("click", () => {
    input.value = "";
    update();
    input.focus();
  });
  update();
})();
</script>'''


def category_page(rows: list[dict[str, str]], category: SubjectCategory) -> str:
    canonical = legacy.absolute(f"/과목별학원/{category.slug}/")
    description = f"전국 371개 지역의 {category.label} 학습 안내입니다. {category.grade_label} {category.subject} 진단, 복습, 오답 관리와 상담 확인 항목을 지역별로 살펴보세요."
    items = [
        {"@type": "ListItem", "position": index, "name": f"{row['근처 수업가능 동네']} {category.label}", "url": legacy.absolute(f"/과목별학원/{category.slug}/{legacy.slug_local(row['근처 수업가능 동네'])}/")}
        for index, row in enumerate(rows, 1)
    ]
    graph = [
        {"@type": "CollectionPage", "@id": canonical + "#webpage", "url": canonical, "name": category.label, "description": description, "inLanguage": "ko-KR", "hasPart": {"@id": canonical + "#regions"}},
        {"@type": "BreadcrumbList", "@id": canonical + "#breadcrumb", "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "홈", "item": legacy.BASE_URL + "/"},
            {"@type": "ListItem", "position": 2, "name": "과목별학원", "item": legacy.absolute("/과목별학원/")},
            {"@type": "ListItem", "position": 3, "name": category.label, "item": canonical},
        ]},
        {"@type": "ItemList", "@id": canonical + "#regions", "name": f"{category.label} 지역 목록", "numberOfItems": len(items), "itemListElement": items},
    ]
    body = f'''{legacy.nav("과목별학원")}
  <main><section class="page-hero"><nav class="breadcrumb" aria-label="현재 위치"><a href="/">홈</a><span aria-hidden="true">/</span><a href="/과목별학원/">과목별학원</a><span aria-hidden="true">/</span><span aria-current="page">{legacy.esc(category.label)}</span></nav><p class="eyebrow">SUBJECT ACADEMY DIRECTORY</p><h1>{legacy.esc(category.label)}</h1><p class="lead">{legacy.esc(description)}</p><div class="subject-count"><strong>{len(rows)}</strong><span>지역별 학습 안내</span></div></section>
  <section class="section subject-directory"><div class="section-head"><p class="eyebrow">지역 찾기</p><h2>광역 지역과 시·군·구 순서로 찾기</h2><p class="lead">지역을 선택한 뒤 동네별 {legacy.esc(category.label)} 상담 안내를 확인할 수 있습니다.</p></div><div class="subject-search" data-subject-search><label for="subject-local-search-{legacy.esc(category.slug)}">지역 검색</label><div class="subject-search-controls"><input id="subject-local-search-{legacy.esc(category.slug)}" data-local-search-input type="search" placeholder="동네 또는 지역을 검색하세요" autocomplete="off"><button data-local-search-reset type="button" disabled>초기화</button></div><p data-local-search-status class="subject-search-status" aria-live="polite">전체 {len(rows)}개 지역</p></div>{region_directory(rows, category)}</section></main>
{legacy.footer()}
{DIRECTORY_SEARCH_SCRIPT}'''
    return legacy.shell(legacy.head_html(category.label, description, canonical, "/assets/generated/coaching-center-hero-v2.png", graph, og_type="website"), body)


def subject_root_page() -> str:
    canonical = legacy.absolute("/과목별학원/")
    description = "학년과 과목별로 전국 지역 학습 안내를 찾는 채움학습 과목별학원 허브입니다. 진단, 복습, 오답 관리와 상담 기준을 확인하세요."
    cards = "".join(
        f'<a class="subject-category-card" href="/과목별학원/{category.slug}/"><span>{index:02d}</span><strong>{legacy.esc(category.label)}</strong><small>371개 지역별 안내 보기</small></a>'
        for index, category in enumerate(SUBJECT_CATALOG, 1)
    )
    items = [
        {"@type": "ListItem", "position": index, "name": category.label, "url": legacy.absolute(f"/과목별학원/{category.slug}/")}
        for index, category in enumerate(SUBJECT_CATALOG, 1)
    ]
    graph = [
        {"@type": "CollectionPage", "@id": canonical + "#webpage", "url": canonical, "name": "과목별학원", "description": description, "inLanguage": "ko-KR", "hasPart": {"@id": canonical + "#categories"}},
        {"@type": "BreadcrumbList", "@id": canonical + "#breadcrumb", "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "홈", "item": legacy.BASE_URL + "/"},
            {"@type": "ListItem", "position": 2, "name": "과목별학원", "item": canonical},
        ]},
        {"@type": "ItemList", "@id": canonical + "#categories", "name": "과목별학원 카테고리", "numberOfItems": len(items), "itemListElement": items},
    ]
    body = f'''{legacy.nav("과목별학원")}
  <main><section class="page-hero"><nav class="breadcrumb" aria-label="현재 위치"><a href="/">홈</a><span aria-hidden="true">/</span><span aria-current="page">과목별학원</span></nav><p class="eyebrow">SUBJECT ACADEMY HUB</p><h1>과목별학원</h1><p class="lead">{legacy.esc(description)}</p></section>
  <section class="section"><div class="section-head"><p class="eyebrow">학년·과목 선택</p><h2>필요한 학습 안내부터 확인하세요</h2><p class="lead">초등 3학년부터 고등 2학년까지 학년과 과목을 선택한 뒤 지역별 상담 기준을 살펴볼 수 있습니다.</p></div><div class="subject-category-grid">{cards}</div></section></main>
{legacy.footer()}'''
    return legacy.shell(legacy.head_html("과목별학원", description, canonical, "/assets/generated/coaching-center-hero-v2.png", graph, og_type="website"), body)


def sentence_units(value: str) -> set[str]:
    return {normalized(part) for part in re.split(r"(?<=[.!?。])\s+|[\r\n]+", re.sub(r"<[^>]+>", " ", value)) if len(normalized(part)) >= 24}


def word_shingles(value: str, size: int = 12) -> set[tuple[str, ...]]:
    words = re.findall(r"[가-힣A-Za-z0-9]+", re.sub(r"<[^>]+>", " ", value).lower())
    return {tuple(words[index:index + size]) for index in range(max(0, len(words) - size + 1))}


def mask_for_similarity(copy: PageCopy, row: dict[str, str], category: SubjectCategory) -> str:
    value = " ".join(copy.audit_strings()).lower()
    masks = [
        copy.title, row.get("근처 수업가능 동네", ""), row.get("지역", ""), row.get("시or구", ""),
        row.get("센터명", ""), category.label, category.grade_label, category.grade_code,
        *split_values(row.get(category.school_field, "")),
    ]
    for mask in sorted(filter(None, masks), key=len, reverse=True):
        value = value.replace(mask.lower(), " <entity> ")
    return re.sub(r"\d+", "<n>", value)


def assert_masked_jaccard_below(documents: list[set[tuple[str, ...]]], threshold: float) -> float:
    # Exact threshold guard using the standard global-order prefix filter.  A
    # pair at or above the threshold must share a prefix token; candidates are
    # then measured with exact Jaccard rather than an approximation.
    frequencies = Counter(token for document in documents for token in document)
    ordered = [sorted(document, key=lambda token: (frequencies[token], token)) for document in documents]
    inverted: dict[tuple[str, ...], list[int]] = defaultdict(list)
    maximum = 0.0
    for index, tokens in enumerate(ordered):
        prefix_length = max(1, len(tokens) - math.ceil(threshold * len(tokens)) + 1)
        candidates: set[int] = set()
        for token in tokens[:prefix_length]:
            candidates.update(inverted[token])
        for other in candidates:
            low, high = len(documents[other]), len(documents[index])
            if min(low, high) / max(low, high) < threshold:
                continue
            first, second = documents[other], documents[index]
            intersection = len(first & second)
            score = intersection / (low + high - intersection)
            maximum = max(maximum, score)
            if score >= threshold:
                raise RuntimeError(f"masked Jaccard {score:.6f} >= {threshold}: documents {other}, {index}")
        for token in tokens[:prefix_length]:
            inverted[token].append(index)
    return maximum


def preflight(
    rows: list[dict[str, str]], categories: Sequence[SubjectCategory],
    source_cells: dict[str, list[str]], copies: dict[tuple[str, int], PageCopy],
) -> dict[str, object]:
    expected = len(categories) * EXPECTED_ROWS
    if len(copies) != expected:
        raise RuntimeError(f"생성 상세 페이지 수 불일치: {len(copies)} != {expected}")
    all_paragraphs: set[str] = set()
    all_sections: set[str] = set()
    all_documents: set[str] = set()
    masked_documents: list[set[tuple[str, ...]]] = []
    source_sentence_reuse = 0
    source_shingle_reuse = 0
    terminal_missing = 0
    question_endings_checked = 0
    long_action_reuse = 0
    known_pattern_counts: Counter[str] = Counter()
    for category in categories:
        raw_cells = source_cells[category.slug]
        if len(raw_cells) != EXPECTED_ROWS:
            raise RuntimeError(f"원고 행 수 불일치: {category.slug}")
        for row_index, row in enumerate(rows):
            copy = copies[(category.slug, row_index)]
            strings = [normalized(value) for value in copy.audit_strings() if normalized(value)]
            if len(strings) != len(set(strings)):
                raise RuntimeError(f"페이지 내부 문구 중복: {category.slug}/{row_index + 1}")
            paragraphs = [normalized(paragraph) for _, values in copy.sections for paragraph in values]
            paragraphs += [normalized(value) for value in copy.checklist]
            paragraphs += [normalized(value) for _, value in copy.scenarios]
            paragraphs += [normalized(answer) for _, answer in copy.faqs]
            duplicate_paragraph = next((value for value in paragraphs if value in all_paragraphs), None)
            if duplicate_paragraph:
                raise RuntimeError(f"페이지 간 동일 문단: {category.slug}/{row_index + 1}: {duplicate_paragraph[:80]}")
            all_paragraphs.update(paragraphs)
            section_texts = [normalized(heading + " " + " ".join(values)) for heading, values in copy.sections]
            section_texts += [normalized(" ".join(copy.checklist)), normalized(" ".join(value for _, value in copy.scenarios))]
            duplicate_section = next((value for value in section_texts if value in all_sections), None)
            if duplicate_section:
                raise RuntimeError(f"페이지 간 동일 섹션: {category.slug}/{row_index + 1}")
            all_sections.update(section_texts)
            document = normalized(" ".join(strings))
            public_complete = [copy.description, copy.summary]
            for _, values in copy.sections:
                public_complete.extend(values)
            for _, paragraph in copy.scenarios:
                public_complete.append(paragraph)
            public_complete.extend(answer for _, answer in copy.faqs)
            missing_here = [value for value in public_complete if not re.search(r"[.!?。！？][\"'’”)]*$", value)]
            question_complete = [*copy.checklist, *(question for question, _ in copy.faqs)]
            question_endings_checked += len(question_complete)
            question_missing = [value for value in question_complete if not re.search(r"\?[\"'’”)]*$", value)]
            terminal_missing += len(missing_here) + len(question_missing)
            if missing_here or question_missing:
                sample = (missing_here or question_missing)[0]
                raise RuntimeError(
                    f"공개 완성문 종결부호 누락: {category.slug}/{row_index + 1}: {sample[:80]}"
                )
            for name, pattern in KNOWN_COPY_PATTERNS.items():
                known_pattern_counts[name] += len(pattern.findall(document))
            blocked_pattern = next((name for name, count in known_pattern_counts.items() if count), None)
            if blocked_pattern:
                raise RuntimeError(f"known copy pattern 검출: {blocked_pattern}: {known_pattern_counts[blocked_pattern]}")
            for action in (*PLAN_ACTIONS[category.subject], *FEEDBACKS):
                occurrences = document.count(action)
                if occurrences > 1:
                    long_action_reuse += occurrences - 1
                    raise RuntimeError(
                        f"본문/FAQ 긴 행동문 재사용: {category.slug}/{row_index + 1}: {action[:80]}"
                    )
            forbidden = next((phrase for phrase in FORBIDDEN_AVAILABILITY_PHRASES if phrase in document), None)
            if forbidden:
                raise RuntimeError(f"학년 가능 여부 금지 표현: {category.slug}/{row_index + 1}: {forbidden}")
            if document in all_documents:
                raise RuntimeError(f"페이지 간 동일 문서: {category.slug}/{row_index + 1}")
            all_documents.add(document)
            generated_sentences = sentence_units(" ".join(paragraphs))
            source_sentences = sentence_units(raw_cells[row_index])
            overlap = generated_sentences & source_sentences
            if overlap:
                source_sentence_reuse += len(overlap)
                raise RuntimeError(f"원고 문장 재사용: {category.slug}/{row_index + 1}: {next(iter(overlap))[:80]}")
            generated_shingles = word_shingles(" ".join(paragraphs))
            source_shingles = word_shingles(raw_cells[row_index])
            shingle_overlap = generated_shingles & source_shingles
            if shingle_overlap:
                source_shingle_reuse += len(shingle_overlap)
                raise RuntimeError(f"원고 12단어 shingle 재사용: {category.slug}/{row_index + 1}")
            masked_documents.append(word_shingles(mask_for_similarity(copy, row, category), size=5))
    masked_max = assert_masked_jaccard_below(masked_documents, MASKED_JACCARD_LIMIT)
    return {
        "details": expected,
        "paragraphs": len(all_paragraphs),
        "sections": len(all_sections),
        "documents": len(all_documents),
        "source_exact_sentence_reuse": source_sentence_reuse,
        "source_12word_shingle_reuse": source_shingle_reuse,
        "within_page_duplicates": 0,
        "terminal_punctuation_missing": terminal_missing,
        "question_endings_checked": question_endings_checked,
        "body_faq_long_action_reuse": long_action_reuse,
        "known_copy_pattern_violations": sum(known_pattern_counts.values()),
        **{f"known_{name}": count for name, count in sorted(known_pattern_counts.items())},
        "masked_candidate_max_jaccard": round(masked_max, 6),
    }


def write_page(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8", newline="\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="채움학습 미생성 학년·과목 페이지 생성")
    selection = parser.add_mutually_exclusive_group()
    selection.add_argument("--category", choices=[category.slug for category in NEW_CATEGORIES])
    selection.add_argument(
        "--hubs-only", action="store_true",
        help="전체 preflight 후 신규 14개 허브와 과목별학원 루트만 갱신",
    )
    parser.add_argument("--check-only", action="store_true", help="입력·카피·중복 사전검사만 하고 파일은 쓰지 않음")
    args = parser.parse_args()
    write_categories = (BY_SLUG[args.category],) if args.category else NEW_CATEGORIES
    # A partial write is never allowed to weaken release validation: all 5,194
    # planned detail pages and all fourteen image permutations are preflighted.
    validation_categories = NEW_CATEGORIES
    rows = load_centers()
    images = representative_images()
    representative_assignments, representative_report = build_representative_assignments(
        rows, images, validation_categories,
    )
    source_cells = {category.slug: load_source_cells(category) for category in validation_categories}
    signals: dict[tuple[str, int], SourceSignal] = {}
    copies: dict[tuple[str, int], PageCopy] = {}
    signal_stats: Counter[str] = Counter()
    for category in validation_categories:
        for row_index, row in enumerate(rows):
            signal = extract_signal(source_cells[category.slug][row_index], category, row_index)
            signals[(category.slug, row_index)] = signal
            copies[(category.slug, row_index)] = build_copy(row, category, signal, row_index)
            signal_stats["source_cells"] += 1
            signal_stats["source_error_or_empty_fallbacks"] += int(not signal.source_ok)
            signal_stats["primary_positive_evidence"] += int(signal.primary_positive_evidence)
            signal_stats["secondary_positive_evidence"] += int(signal.secondary_positive_evidence)
            signal_stats["template_fallback_assignments"] += int(signal.fallback_key is not None)
            if signal.fallback_key:
                signal_stats[signal.fallback_key] += 1
    report = preflight(rows, validation_categories, source_cells, copies)
    report.update(representative_report)
    report.update(signal_stats)
    report["validated_categories"] = len(validation_categories)
    report["selected_write_categories"] = [category.slug for category in write_categories]
    report["write_mode"] = "hubs-only" if args.hubs_only else "full"
    if args.check_only:
        print(report)
        return
    before = protected_hashes()
    detail_before = new_detail_hashes() if args.hubs_only else None
    for category in write_categories:
        category_root = SITE / "과목별학원" / category.slug
        if category.existing or category_root.name in {item.slug for item in PROTECTED_CATEGORIES}:
            raise RuntimeError(f"보호 범주 쓰기 차단: {category.slug}")
        write_page(category_root / "index.html", category_page(rows, category))
        if not args.hubs_only:
            for row_index, row in enumerate(rows):
                slug = legacy.slug_local(row["근처 수업가능 동네"])
                copy = copies[(category.slug, row_index)]
                signal = signals[(category.slug, row_index)]
                write_page(
                    category_root / slug / "index.html",
                    local_page(row, row_index, category, copy, signal, rows, representative_assignments),
                )
    write_page(SITE / "과목별학원" / "index.html", subject_root_page())
    assert_protected_unchanged(before)
    if detail_before is not None:
        assert_new_details_unchanged(detail_before)
    generated = 0 if args.hubs_only else len(write_categories) * EXPECTED_ROWS
    if not args.hubs_only and len(write_categories) == len(NEW_CATEGORIES) and generated != EXPECTED_NEW_DETAILS:
        raise RuntimeError(f"전체 신규 상세 페이지 수 불일치: {generated}")
    print({
        **report, "written_categories": len(write_categories), "hubs": len(write_categories),
        "subject_root": 1, "written_details": generated,
        "protected_categories": "byte-identical",
        "new_details": "byte-identical" if args.hubs_only else "regenerated",
    })


if __name__ == "__main__":
    main()
